# -*- coding: utf-8 -*-
"""Excel generator — builds a fresh, high-end RTL workbook from scratch every run.

Workbook (same DNA at monthly / quarterly / annual):
  * "סקירה" (overview): KPI cards -> two income pictures side by side ->
    clean trend/breakdown charts -> targets attainment -> analytics -> missing report.
  * Row-level, period-scoped detail sheets: פרויקטים · לידים · הכנסות · שתפים · משימות
    (accent header, zebra rows, frozen header, autofilter).
  * "נתונים" (hidden): the chart source.

The code computes every number; this module only lays it out. Open-schema: any new
dimension value (source, category, employee, status) appears automatically as a new
row/column in the existing style. Style tokens live ONLY in style_config.py.
"""
import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.utils import get_column_letter
from openpyxl.chart import LineChart, BarChart, Reference
from openpyxl.chart.shapes import GraphicalProperties
from openpyxl.chart.label import DataLabelList
from openpyxl.chart.text import RichText
from openpyxl.drawing.image import Image
from openpyxl.drawing.line import LineProperties
from openpyxl.drawing.text import Paragraph, ParagraphProperties, CharacterProperties
from openpyxl.formatting.rule import DataBarRule
from openpyxl.styles import Border

from scripts import style_config as sc
from scripts import targets as tg
from scripts import viewmodel as vm

STUDIO = sc.STUDIO_NAME
DATA_SHEET = "נתונים"          # hidden chart-source sheet
NCOLS = 24                      # overview working-grid width (uniform columns)
LOGO_COL = NCOLS                # header logo anchored at the LEFT corner (RTL end side)
TREND_COLORS = [sc.PRIMARY, sc.SECONDARY, sc.MUTED]  # ≤2 accents + gray

# semantic-token -> hex resolvers for view-model blocks (the view-model stays palette-free)
_HEADER_HEX = {"pic_a": sc.PIC_A, "pic_b": sc.PIC_B, "primary": sc.PRIMARY}
_SEM_HEX = {"good": sc.GOOD, "bad": sc.BAD, "warn": sc.WARN, "primary": sc.PRIMARY, None: sc.TEXT}
_FMT_XL = {"money": sc.FMT_MONEY, "int": sc.FMT_INT, "pct": sc.FMT_PCT, "days": sc.FMT_DAYS}
# data-label number formats (values rendered directly on the charts)
_DLBL_MONEY = '"₪ "#,##0'
_DLBL_INT = '#,##0'


# ============================================================== low-level cell
def _cell(ws, r, c, value=None, fmt=None, fnt=None, fillc=None, align=None):
    cell = ws.cell(row=r, column=c, value=value)
    cell.font = fnt or sc.font()
    if fillc:
        cell.fill = sc.fill(fillc)
    if align is not None:
        cell.alignment = align
    if fmt:
        cell.number_format = fmt
    return cell


def _fill_range(ws, r1, c1, r2, c2, hexc):
    f = sc.fill(hexc)
    for r in range(r1, r2 + 1):
        for c in range(c1, c2 + 1):
            ws.cell(row=r, column=c).fill = f


def _add_side(cell, which, side):
    b = cell.border
    kw = {"left": b.left, "right": b.right, "top": b.top, "bottom": b.bottom}
    kw[which] = side
    cell.border = Border(**kw)


def _box(ws, r1, c1, r2, c2, top=None, bottom=None, left=None, right=None):
    for c in range(c1, c2 + 1):
        if top:
            _add_side(ws.cell(r1, c), "top", top)
        if bottom:
            _add_side(ws.cell(r2, c), "bottom", bottom)
    for r in range(r1, r2 + 1):
        if left:
            _add_side(ws.cell(r, c1), "left", left)
        if right:
            _add_side(ws.cell(r, c2), "right", right)


def _merge(ws, r1, c1, r2, c2):
    ws.merge_cells(start_row=r1, start_column=c1, end_row=r2, end_column=c2)


def _prep_sheet(ws):
    ws.sheet_view.rightToLeft = True
    ws.sheet_view.showGridLines = False


def _section(ws, r, title, last_col=NCOLS):
    """Full-width accent section band (white bold text on PRIMARY)."""
    _merge(ws, r, 1, r, last_col)
    _cell(ws, r, 1, title, fnt=sc.font(sc.SZ_SECTION, True, sc.WHITE),
          fillc=sc.PRIMARY, align=sc.right())
    ws.row_dimensions[r].height = 24
    return r + 1


# ============================================================== display helpers
def _m(v):   # money string for label/sub text (zero -> em dash)
    if v is None or round(v) == 0:
        return "—"
    return "₪ {:,.0f}".format(v)


def _i(v):
    if v is None or round(v) == 0:
        return "—"
    return "{:,.0f}".format(v)


def _p(v):
    return "—" if v is None else "{:.1f}%".format(v * 100)


def _delta_chip(key, cur, prev):
    """Neutral-gray Δ chip vs previous period (status colors stay for targets)."""
    if prev is None or cur is None:
        return ""
    if key == "conversion":
        d = (cur - prev) * 100.0
        if abs(d) < 0.05:
            return "• ללא שינוי"
        return ("▲ " if d > 0 else "▼ ") + sc.lrm("{:.1f} נק׳".format(abs(d)))
    if not prev:
        return ""
    pct = (cur - prev) / prev * 100.0
    if abs(pct) < 0.5:
        return "• ללא שינוי"
    return ("▲ " if pct > 0 else "▼ ") + sc.lrm("{:.0f}%".format(abs(pct)))


# ============================================================== title band
def _logo_image(ws, row, height_px, col=1):
    """Place the studio logo in the overview header, scaled to `height_px` with aspect
    preserved (the wordmark carries the studio name). Returns True if placed. The image is
    NOT a cell value, so cell-by-cell determinism and the test gate are unaffected; a no-op
    when the asset is absent."""
    path = sc.logo_path()
    if not path:
        return False
    try:
        img = Image(path)
        w0, h0 = img.width, img.height
        img.height = height_px
        img.width = round(height_px * (w0 / h0)) if h0 else img.width
        ws.add_image(img, "{}{}".format(get_column_letter(col), row))
        return True
    except Exception:
        return False


def _title(ws, report):
    z = report["zoom"]
    if z == "quarterly":
        token = sc.lrm("Q{} {}".format(report["quarter"], report["year"]))
        t = "דשבורד רבעוני — " + token
    elif z == "annual":
        t = "דשבורד שנתי — " + sc.lrm(report["year"])
    else:
        t = "דשבורד חודשי — " + sc.heb_month(report["months"][0])

    _merge(ws, 1, 1, 1, NCOLS)
    _cell(ws, 1, 1, t, fnt=sc.font(sc.SZ_TITLE, True, sc.PRIMARY_DK), align=sc.right())
    ws.row_dimensions[1].height = 32

    half = NCOLS // 2
    # row 2 — studio logo on the LEFT (end side in RTL); date-of-issue on the right.
    today = datetime.date.today().strftime("%d/%m/%Y")
    _merge(ws, 2, 1, 2, half)
    _cell(ws, 2, 1, sc.S_GENERATED + ": " + sc.lrm(today),
          fnt=sc.font(sc.SZ_SUBTITLE, False, sc.MUTED), align=sc.right())
    placed = _logo_image(ws, row=2, height_px=46, col=LOGO_COL)
    ws.row_dimensions[2].height = 52 if placed else 18
    if not placed:                       # fallback: name text on the same (left) side
        _merge(ws, 2, half + 1, 2, NCOLS)
        _cell(ws, 2, half + 1, STUDIO, fnt=sc.font(sc.SZ_SUBTITLE, False, sc.MUTED), align=sc.left())

    # thin accent rule
    _merge(ws, 3, 1, 3, NCOLS)
    _fill_range(ws, 3, 1, 3, NCOLS, sc.PRIMARY)
    ws.row_dimensions[3].height = 4
    ws.row_dimensions[4].height = 22     # breathing room before the KPI cards
    return 5


# ============================================================== KPI cards
def _card(ws, top, start, span, value, fmt, label, sub, value_color=sc.TEXT):
    c2 = start + span - 1
    _fill_range(ws, top, start, top + 2, c2, sc.CARD)
    # value (numeric -> number_format renders zero as —)
    _merge(ws, top, start, top, c2)
    _cell(ws, top, start, value, fmt=fmt,
          fnt=sc.font(sc.SZ_KPI, True, value_color), align=sc.center(wrap=False))
    ws.row_dimensions[top].height = 34
    # label
    _merge(ws, top + 1, start, top + 1, c2)
    _cell(ws, top + 1, start, label, fnt=sc.font(sc.SZ_KPILABEL, False, sc.MUTED),
          align=sc.center(wrap=False))
    ws.row_dimensions[top + 1].height = 16
    # sub (Δ chip or value-of-deals)
    _merge(ws, top + 2, start, top + 2, c2)
    _cell(ws, top + 2, start, sub, fnt=sc.font(sc.SZ_DELTA, False, sc.MUTED),
          align=sc.center(wrap=False))
    ws.row_dimensions[top + 2].height = 16
    # 3px accent top border + subtle box
    _box(ws, top, start, top + 2, c2,
         top=sc.side_thick(sc.PRIMARY), bottom=sc.side_thin(), left=sc.side_thin(), right=sc.side_thin())


def _kpi_cards(ws, report, top):
    prev = report.get("prev") or {}
    L, P, C = report["leads"], report["pipeline"], report["cash"]
    Pr, T, K = report["profit"], report["tasks"], report["cohort"]

    # (value, fmt, label, sub, value_color)  — order = rightmost first (RTL)
    deals_sub = "שווי " + _m(P["value_total"])
    dchip = _delta_chip("deals", P["deals_total"], prev.get("deals"))
    if dchip:
        deals_sub += "   " + dchip
    cards = [
        (C["net_total"], sc.FMT_MONEY, sc.K_NET,
         _delta_chip("net", C["net_total"], prev.get("net")), sc.TEXT),
        (Pr["total"], sc.FMT_MONEY, sc.K_PROFIT,
         _delta_chip("profit", Pr["total"], prev.get("profit")),
         (sc.GOOD if Pr["total"] >= 0 else sc.BAD)),
        (P["deals_total"], sc.FMT_INT, sc.K_DEALS, deals_sub, sc.TEXT),
        (L["total"], sc.FMT_INT, sc.K_LEADS,
         _delta_chip("leads", L["total"], prev.get("leads")), sc.TEXT),
        (K["headline_rate"], sc.FMT_PCT, sc.K_CONV,
         _delta_chip("conversion", K["headline_rate"], prev.get("conversion")), sc.TEXT),
        (T["total"], sc.FMT_INT, sc.K_TASKS,
         _delta_chip("tasks", T["total"], prev.get("tasks")), sc.TEXT),
    ]
    span, gap = 3, 1
    note = "מדדי מפתח" + ("   ·   Δ מול התקופה הקודמת" if report.get("prev") else "")
    _cell(ws, top - 1, 1, note, fnt=sc.font(sc.SZ_NOTE, False, sc.MUTED), align=sc.right())
    for k, card in enumerate(cards):
        start = 1 + k * (span + gap)
        if start + span - 1 > NCOLS:
            break
        _card(ws, top, start, span, *card)
    return top + 4  # 3 card rows + 1 spacer


# ============================================================== income pictures
def _value_cols(report):
    """Month/summary columns — delegates to the view-model (single home for the rule)."""
    return vm.value_columns(report)


def _put_span(ws, r, c, span, value, fnt, fillc, align, fmt=None):
    """Write one logical cell anchored at grid col `c`, occupying `span` grid columns.
    When span>1 the columns are merged and the fill painted across the whole span (so a
    wide label/value isn't clipped by the uniform 12-unit grid). span==1 → a plain cell,
    identical to a direct `_cell`, so default-span callers stay byte-identical."""
    if span > 1:
        _merge(ws, r, c, r, c + span - 1)
        if fillc:
            _fill_range(ws, r, c, r, c + span - 1, fillc)
    _cell(ws, r, c, value, fmt=fmt, fnt=fnt, fillc=fillc, align=align)
    return c + span


def _matrix_block(ws, top, start, block, vcols, draw_title=True, label_span=1, val_span=1):
    """Render any monthly+summary view-model Block (income pictures, expenses/profit,
    attribution, source, performer). `draw_title` colors a title band in the block's
    header color; pass False when a full-width `_section` band already precedes it.
    Per-row color (e.g. profit sign) comes from the Row's semantic token. `label_span`/
    `val_span` let each logical column occupy several merged grid columns so wide content
    — e.g. the income-picture titles/labels — fits without wrapping; both default to 1,
    leaving the analytics matrices byte-identical."""
    width = label_span + len(vcols) * val_span
    end = start + width - 1
    r = top
    if draw_title:
        _merge(ws, r, start, r, end)
        _cell(ws, r, start, block.title, fnt=sc.font(sc.SZ_SECTION, True, sc.WHITE),
              fillc=_HEADER_HEX.get(block.color, sc.PRIMARY), align=sc.right())
        ws.row_dimensions[r].height = 22
        r += 1
    # header row
    hr = r
    c = _put_span(ws, hr, start, label_span, sc.S_METRIC,
                  sc.font(sc.SZ_HEAD, True), sc.HEADER_FILL, sc.right())
    for kind, m in vcols:
        lab = sc.S_SUMMARY if kind == "sum" else sc.heb_month(m)
        c = _put_span(ws, hr, c, val_span, lab, sc.font(sc.SZ_HEAD, True), sc.HEADER_FILL, sc.center())
    # data rows
    r = hr + 1
    for i, row in enumerate(block.rows):
        band = sc.BAND if i % 2 == 1 else None
        col = _SEM_HEX.get(row.color, sc.TEXT)
        lab_fnt = sc.font(sc.SZ_BODY, True, col)
        val_fnt = sc.font(sc.SZ_BODY, True, col) if row.color else None
        fmt = _FMT_XL[row.fmt]
        c = _put_span(ws, r, start, label_span, row.label, lab_fnt, band, sc.right())
        for kind, m in vcols:
            v = row.summary if kind == "sum" else (row.by_month.get(m) if row.by_month else None)
            c = _put_span(ws, r, c, val_span, v, val_fnt, band, sc.center(), fmt=fmt)
        if label_span > 1 or val_span > 1:        # spaced income rows; analytics keep default height
            ws.row_dimensions[r].height = 19
        r += 1
    _box(ws, hr, start, r - 1, end, bottom=sc.side_thin())
    return r


def _income_layout(n_vcols):
    """Grid-column spans for the income pictures so each block fills ~half the overview
    width — title on one line, labels/values padded, nothing clipped — and both blocks
    come out equal width with a clean central gutter. Adapts to the column count
    (monthly: one wide value column; quarterly: four narrower month+summary columns)."""
    gutter = 2
    block_w = (NCOLS - gutter) // 2                  # target grid-cols per block (= 11)
    val_span = 2 if n_vcols >= 3 else 4              # wider value cells when few columns
    label_span = max(3, block_w - n_vcols * val_span)
    return label_span, val_span


def _income_side_by_side(ws, report, top):
    a, b = vm.income_blocks(report)
    vcols = _value_cols(report)
    label_span, val_span = _income_layout(len(vcols))
    width = label_span + len(vcols) * val_span
    # Picture A on the right (start col 1); Picture B right-justified to the left edge so
    # the two are equal width with a clean central gutter between them. (RTL: col 1 = right.)
    b_start = NCOLS - width + 1
    bottom_a = _matrix_block(ws, top, 1, a, vcols, draw_title=True,
                             label_span=label_span, val_span=val_span)
    bottom_b = _matrix_block(ws, top, b_start, b, vcols, draw_title=True,
                             label_span=label_span, val_span=val_span)
    return max(bottom_a, bottom_b) + 1


# ============================================================== chart source + charts
def _chart_data_sheet(wb, report):
    ds = wb.create_sheet(DATA_SHEET)
    _prep_sheet(ds)
    cs = vm.chart_series(report)          # one source for the Excel data sheet + the HTML SVG
    months = cs["months"]
    # Short series names — these become the chart legend entries (compact, no overlap).
    heads = ["חודש", "נטו", "ברוטו", "שווי", "פניות", "עסקאות", "משימות", "רווח"]
    for c, h in enumerate(heads, 1):
        _cell(ds, 1, c, h, fnt=sc.font(10, True), fillc=sc.HEADER_FILL, align=sc.center())
    series_cols = [("net", sc.FMT_MONEY), ("gross", sc.FMT_MONEY), ("value", sc.FMT_MONEY),
                   ("leads", sc.FMT_INT), ("deals", sc.FMT_INT), ("tasks", sc.FMT_INT),
                   ("profit", sc.FMT_MONEY)]
    for i, m in enumerate(months):
        rr = 2 + i
        _cell(ds, rr, 1, sc.heb_month(m), align=sc.center())
        for ci, (key, fmt) in enumerate(series_cols):
            _cell(ds, rr, 2 + ci, cs[key][i], fmt=fmt, align=sc.center())
    last = 1 + len(months)

    def _kv(start_row, title, pairs, fmt):
        _cell(ds, start_row, 1, title, fnt=sc.font(10, True), fillc=sc.SUBHEAD_FILL)
        _cell(ds, start_row, 2, sc.S_VALUE, fnt=sc.font(10, True), fillc=sc.SUBHEAD_FILL, align=sc.center())
        rr = start_row + 1
        for k, v in pairs:                # pre-sorted in the view-model (stable tiebreak)
            _cell(ds, rr, 1, str(k))
            _cell(ds, rr, 2, v, fmt=fmt, align=sc.center())
            rr += 1
        return start_row + 1, rr - 1

    row = last + 3
    tables = {}
    tables["src"] = _kv(row, sc.M_LEADS + " — מקור", cs["src"], sc.FMT_INT); row += len(cs["src"]) + 3
    tables["perf"] = _kv(row, sc.M_TASKS + " — מבצע", cs["perf"], sc.FMT_INT); row += len(cs["perf"]) + 3
    tables["cat"] = _kv(row, sc.M_EXP + " — קטגוריה", cs["cat"], sc.FMT_MONEY)
    ds.column_dimensions["A"].width = 26
    for c in range(2, 9):
        ds.column_dimensions[get_column_letter(c)].width = 13
    ds.sheet_state = "hidden"
    return ds, tables


def _clean_axes(chart):
    chart.y_axis.majorGridlines = None
    chart.x_axis.majorGridlines = None
    chart.x_axis.delete = False
    chart.y_axis.delete = False


def _style_lines(chart):
    for i, s in enumerate(chart.series):
        col = TREND_COLORS[i % len(TREND_COLORS)]
        gp = GraphicalProperties()
        gp.line = LineProperties(solidFill=col, w=28575)  # ~2.25pt
        s.graphicalProperties = gp
        s.smooth = False


def _style_bars(chart, color=sc.PRIMARY):
    for s in chart.series:
        s.graphicalProperties = GraphicalProperties(solidFill=color)


def _data_labels(chart, numfmt, position=None):
    """Render each point's value directly on the chart (clean: small 8pt, formatted)."""
    dl = DataLabelList()
    dl.showVal = True
    dl.showSerName = False
    dl.showCatName = False
    dl.showLegendKey = False
    dl.showPercent = False
    dl.showBubbleSize = False
    dl.numFmt = numfmt
    if position:
        dl.dLblPos = position             # bars: "outEnd"; lines: "t"
    cp = CharacterProperties(sz=800)      # 8pt (sz is in 1/100 pt)
    dl.txPr = RichText(p=[Paragraph(pPr=ParagraphProperties(defRPr=cp), endParaRPr=cp)])
    chart.dataLabels = dl


def _axis_ceiling(v):
    """Smallest 'nice' axis max (mantissa 1·2·2.5·5·10) ≥ v·1.15 — leaves headroom so an
    end-of-bar data label never clips the plot edge. Deterministic, presentation-only."""
    if v is None or v <= 0:
        return 1
    target = v * 1.15
    mag = 1
    while mag * 10 <= target:
        mag *= 10
    for m in (1, 2, 2.5, 5, 10):
        top = m * mag
        if top >= target:
            return int(top) if top == int(top) else top
    return 10 * mag


def _add_charts(ws, ds, tables, report, anchor):
    n = len(report["months"])
    cats = Reference(ds, min_col=1, min_row=2, max_row=1 + n)
    charts = []

    if n > 1:
        money = LineChart()
        money.title = "מגמת הכנסה ושווי (₪)"
        money.add_data(Reference(ds, min_col=2, max_col=4, min_row=1, max_row=1 + n), titles_from_data=True)
        money.set_categories(cats)
        money.y_axis.scaling.min = 0
        _style_lines(money)
        _data_labels(money, _DLBL_MONEY, "t")
        charts.append(money)

        counts = LineChart()
        counts.title = "מגמת פניות / עסקאות / משימות"
        counts.add_data(Reference(ds, min_col=5, max_col=7, min_row=1, max_row=1 + n), titles_from_data=True)
        counts.set_categories(cats)
        counts.y_axis.scaling.min = 0
        _style_lines(counts)
        _data_labels(counts, _DLBL_INT, "t")
        charts.append(counts)

        # NEW: monthly profit trend (no y-axis floor — profit may be negative).
        profit = LineChart()
        profit.title = "מגמת רווח חודשי (₪)"
        profit.add_data(Reference(ds, min_col=8, min_row=1, max_row=1 + n), titles_from_data=True)
        profit.set_categories(cats)
        _style_lines(profit)
        _data_labels(profit, _DLBL_MONEY, "t")
        charts.append(profit)
    else:
        # Single month: a HORIZONTAL 3-bar composition (net / gross / value) — mirrors the
        # HTML "הרכב ההכנסה" view. Categories come from the series headers and the single
        # data row is the one series, so the labels sit at the bar ends with clear room and
        # never collide with the title or the category axis (the old vertical 3-bar layout
        # crowded the tallest bar's label into the title). Value axis given nice headroom.
        money = BarChart()
        money.type = "bar"
        money.title = "הכנסה ושווי לחודש (₪)"
        money.add_data(Reference(ds, min_col=2, max_col=4, min_row=2, max_row=2),
                       from_rows=True, titles_from_data=False)
        money.set_categories(Reference(ds, min_col=2, max_col=4, min_row=1, max_row=1))
        money.y_axis.scaling.min = 0
        money.y_axis.scaling.max = _axis_ceiling(max(
            report["cash"]["net_total"], report["cash"]["gross_total"],
            report["pipeline"]["value_total"]))
        money.legend = None
        _style_bars(money, sc.PRIMARY)
        _data_labels(money, _DLBL_MONEY, "outEnd")
        charts.append(money)

    def _bar(table, title, numfmt):
        first, last = table
        if last < first:
            return None
        b = BarChart()
        b.type = "bar"
        b.title = title
        b.add_data(Reference(ds, min_col=2, min_row=first, max_row=last), titles_from_data=False)
        b.set_categories(Reference(ds, min_col=1, min_row=first, max_row=last))
        b.legend = None
        _style_bars(b, sc.PRIMARY)
        _data_labels(b, numfmt, "outEnd")
        return b

    for tbl, title, numfmt in ((tables["src"], sc.M_LEADS + " לפי מקור", _DLBL_INT),
                               (tables["perf"], sc.M_TASKS + " לפי מבצע", _DLBL_INT),
                               (tables["cat"], sc.M_EXP + " לפי קטגוריה", _DLBL_MONEY)):
        c = _bar(tbl, title, numfmt)
        if c:
            charts.append(c)

    for ch in charts:
        ch.height, ch.width = 7.6, 15.5
        _clean_axes(ch)
        if ch.legend is not None:
            ch.legend.position = "r"
            ch.legend.overlay = False

    r = anchor
    for i in range(0, len(charts), 2):
        ws.add_chart(charts[i], "A{}".format(r))
        if i + 1 < len(charts):
            ws.add_chart(charts[i + 1], "M{}".format(r))
        r += 17
    return r + 1


# ============================================================== targets
def _targets_block(ws, r, report):
    targets = report.get("targets")
    if not targets:
        return r
    r = _section(ws, r, sc.T_TARGETS)
    heads = [sc.S_METRIC, sc.S_ACTUAL, sc.S_TARGET, sc.S_PROGRESS, "מצב"]
    for j, h in enumerate(heads):
        _cell(ws, r, 1 + j, h, fnt=sc.font(sc.SZ_HEAD, True), fillc=sc.HEADER_FILL,
              align=(sc.right() if j == 0 else sc.center()))
    r += 1
    actuals = tg.actual_for(report)
    first = r
    for key, target in targets.items():
        if key not in actuals:
            continue
        label, actual = actuals[key]
        pct = (actual / target) if (target and actual is not None) else None
        fmt = sc.FMT_PCT if key == "conversion" else (sc.FMT_INT if key in ("leads", "deals", "tasks") else sc.FMT_MONEY)
        band = sc.BAND if (r - first) % 2 == 1 else None
        _cell(ws, r, 1, label, fnt=sc.font(sc.SZ_BODY, True), fillc=band, align=sc.right())
        _cell(ws, r, 2, actual, fmt=fmt, fillc=band, align=sc.center())
        _cell(ws, r, 3, target, fmt=fmt, fillc=band, align=sc.center())
        _cell(ws, r, 4, pct, fmt=sc.FMT_PCT, fillc=band, align=sc.center())
        status = sc.GOOD if (pct is not None and pct >= 1) else (sc.WARN if (pct is not None and pct >= 0.7) else sc.BAD)
        _cell(ws, r, 5, "●", fnt=sc.font(sc.SZ_BODY, True, status), fillc=band, align=sc.center())
        r += 1
    if r > first:
        col = get_column_letter(4)
        rule = DataBarRule(start_type="num", start_value=0, end_type="num", end_value=1.0,
                           color=sc.PRIMARY.replace("#", ""), showValue=True)
        ws.conditional_formatting.add("{c}{a}:{c}{b}".format(c=col, a=first, b=r - 1), rule)
    return r + 1


# ============================================================== analytics
def _kv_table(ws, r, headers, rows):
    for j, h in enumerate(headers):
        _cell(ws, r, 1 + j, h, fnt=sc.font(sc.SZ_HEAD, True), fillc=sc.HEADER_FILL,
              align=(sc.right() if j == 0 else sc.center()))
    r += 1
    first = r
    for i, cols in enumerate(rows):
        band = sc.BAND if i % 2 == 1 else None
        for j, (val, fmt, fnt, align) in enumerate(cols):
            _cell(ws, r, 1 + j, val, fmt=fmt, fnt=fnt, fillc=band,
                  align=(align or (sc.right() if j == 0 else sc.center())))
        r += 1
    _box(ws, first - 1, 1, r - 1, len(headers), bottom=sc.side_thin())
    return r


def _analytics(ws, r, report, data):
    body = sc.font(sc.SZ_BODY)
    bold = sc.font(sc.SZ_BODY, True)
    vcols = _value_cols(report)

    # Expenses & profit — monthly + summary; profit colored by sign.
    r = _section(ws, r, sc.T_EXP)
    r = _matrix_block(ws, r, 1, vm.expenses_block(report), vcols, draw_title=False) + 1

    # Leads by source — monthly + summary (naturally monthly).
    r = _section(ws, r, sc.T_LEADS)
    r = _matrix_block(ws, r, 1, vm.source_matrix(report), vcols, draw_title=False) + 1

    # Tasks by performer (מבצע) — monthly + summary.
    r = _section(ws, r, sc.T_TASKS)
    r = _matrix_block(ws, r, 1, vm.performer_matrix(report), vcols, draw_title=False) + 1

    # Cohort conversion (anchored to arrival month — kept as-is).
    K = report["cohort"]
    r = _section(ws, r, sc.T_COHORT)
    rows = []
    for row in K["rows"]:
        st = (sc.S_MATURING, None, sc.font(sc.SZ_NOTE, False, sc.WARN), None) if row["maturing"] else ("", None, body, None)
        rows.append([
            (sc.heb_month(row["month"]), None, body, None),
            (row["n_leads"], sc.FMT_INT, body, None),
            (row["n_conv"], sc.FMT_INT, body, None),
            (row["rate"], sc.FMT_PCT, body, None),
            st,
        ])
    r = _kv_table(ws, r, ["חודש הגעה", "לידים", "נסגרו", sc.M_CONV, "סטטוס"], rows)
    _cell(ws, r, 1, sc.M_CONV_SETTLED + " (קוהורטות בשלות)", fnt=sc.font(sc.SZ_BODY, True, sc.PRIMARY), align=sc.right())
    _cell(ws, r, 2, K["headline_n"], fmt=sc.FMT_INT, align=sc.center())
    _cell(ws, r, 3, K["headline_conv"], fmt=sc.FMT_INT, align=sc.center())
    _cell(ws, r, 4, K["headline_rate"], fmt=sc.FMT_PCT, fnt=sc.font(sc.SZ_BODY, True, sc.PRIMARY), align=sc.center())
    r += 1
    if K["ttc_median"] is not None:
        _cell(ws, r, 1, sc.M_TTC, fnt=bold, align=sc.right())
        _cell(ws, r, 2, K["ttc_median"], fmt=sc.FMT_DAYS, align=sc.center())
        r += 1
        _merge(ws, r, 1, r, NCOLS)        # time-to-close caveat (biased low until lead history grows)
        _cell(ws, r, 1, "* " + sc.S_TTC_CAVEAT, fnt=sc.font(sc.SZ_NOTE, False, sc.MUTED), align=sc.right())
        r += 1
    r += 1

    # Attribution (informational) — monthly + summary.
    r = _section(ws, r, sc.T_ATTR)
    r = _matrix_block(ws, r, 1, vm.attribution_block(report, data), vcols, draw_title=False)
    _merge(ws, r, 1, r, NCOLS)
    _cell(ws, r, 1, "היתר: הפניה / לקוח חוזר / ישיר / טרם חלון המעקב — אינו חוסר נתונים",
          fnt=sc.font(sc.SZ_NOTE, False, sc.MUTED), align=sc.right())
    r += 2

    # Missing-items report
    r = _section(ws, r, sc.T_MISSING)
    for item in (report.get("missing") or ["אין חוסרים מהותיים"]):
        _merge(ws, r, 1, r, NCOLS)
        _cell(ws, r, 1, "•  " + str(item), fnt=sc.font(sc.SZ_NOTE, False, sc.MUTED), align=sc.right())
        r += 1
    return r


# ============================================================== detail sheets
def _paint_detail_rows(ws, rows, colspecs, hdr):
    """Zebra-paint already-filtered rows (pandas Series or plain dicts) by column kind.
    Returns the first empty row index (for the autofilter range)."""
    r = hdr + 1
    for i, row in enumerate(rows):
        band = sc.BAND if i % 2 == 1 else sc.WHITE
        for j, (key, head, kind, width) in enumerate(colspecs):
            v = row.get(key)
            if kind == "date":
                if v is None or pd.isna(v):
                    v, fmt, align = None, None, sc.center()
                else:
                    v = v.to_pydatetime() if hasattr(v, "to_pydatetime") else v
                    fmt, align = sc.FMT_DATE, sc.center()
            elif kind == "money":
                v = None if (v is None or pd.isna(v)) else float(v)
                fmt, align = sc.FMT_MONEY, sc.center()
            elif kind == "int":
                v = None if (v is None or pd.isna(v)) else float(v)
                fmt, align = sc.FMT_INT, sc.center()
            else:  # text
                v = None if (v is None or (isinstance(v, float) and pd.isna(v))) else str(v)
                fmt, align = None, sc.right()
            _cell(ws, r, 1 + j, v, fmt=fmt, fillc=band, align=align)
        ws.row_dimensions[r].height = 22
        r += 1
    return r


def _detail_sheet(wb, name, colspecs, rows):
    """Row-level detail sheet: accent header (white text), zebra, 22px rows, frozen
    header, autofilter. `rows` are pre-filtered (period-scoped) by the view-model and
    may be pandas Series (the five dimensions) or dicts (expenses replication)."""
    ws = wb.create_sheet(name)
    _prep_sheet(ws)
    ncol = len(colspecs)
    # title band
    _merge(ws, 1, 1, 1, ncol)
    _cell(ws, 1, 1, name, fnt=sc.font(sc.SZ_SECTION + 1, True, sc.PRIMARY_DK), align=sc.right())
    ws.row_dimensions[1].height = 26
    hdr = 3
    # header row (accent, white text)
    for j, (key, head, kind, width) in enumerate(colspecs):
        _cell(ws, hdr, 1 + j, head, fnt=sc.font(sc.SZ_HEAD, True, sc.WHITE), fillc=sc.PRIMARY,
              align=(sc.right() if kind == "text" else sc.center()))
        ws.column_dimensions[get_column_letter(1 + j)].width = width
    ws.row_dimensions[hdr].height = 22

    if not rows:
        _merge(ws, hdr + 1, 1, hdr + 1, ncol)
        _cell(ws, hdr + 1, 1, sc.S_NODATA, fnt=sc.font(sc.SZ_BODY, False, sc.MUTED), align=sc.right())
        ws.row_dimensions[hdr + 1].height = 22
        ws.freeze_panes = "A{}".format(hdr + 1)
        return ws

    r = _paint_detail_rows(ws, rows, colspecs, hdr)
    ws.freeze_panes = "A{}".format(hdr + 1)
    ws.auto_filter.ref = "A{}:{}{}".format(hdr, get_column_letter(ncol), r - 1)
    return ws


def _detail_sheets(wb, report, data):
    months = report["months"]
    for spec in vm.detail_tables(data, months):     # five period-scoped dimensions
        _detail_sheet(wb, spec["name"], spec["cols"], spec["rows"])
    # NEW: expenses detail (monthly-replication; sum of amount rows == overview total).
    _detail_sheet(wb, vm.EXPENSE_SPEC["name"], vm.EXPENSE_SPEC["cols"],
                  vm.expense_detail_rows(data, months))


# ============================================================== main
def build_workbook(report, out_path, data=None):
    wb = Workbook()
    ws = wb.active
    ws.title = sc.T_OVERVIEW
    _prep_sheet(ws)

    # uniform overview grid (wide enough for ₪-formatted values; generous whitespace)
    for c in range(1, NCOLS + 1):
        ws.column_dimensions[get_column_letter(c)].width = 12.0

    r = _title(ws, report)
    r = _kpi_cards(ws, report, r + 1)
    r = _income_side_by_side(ws, report, r)

    # detail sheets (built before charts so the data sheet ends up last)
    if data is not None:
        _detail_sheets(wb, report, data)

    # charts (sourced from the hidden data sheet)
    ds, tables = _chart_data_sheet(wb, report)
    r = _add_charts(ws, ds, tables, report, anchor=r + 1)

    # targets + analytics + missing
    r = _targets_block(ws, r, report)
    r = _analytics(ws, r, report, data)

    wb.save(out_path)
    return out_path
