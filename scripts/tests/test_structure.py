# -*- coding: utf-8 -*-
"""Structural invariants: RTL, chart data labels, frozen+filtered detail sheets, the
two income pictures separate, NO pdf, a clean outputs/ tree, and inline SVG in the HTML."""
import os
import re
import zipfile

import harness
from harness import check, get_ctx, ROOT, _nfc
from openpyxl import load_workbook
from scripts import style_config as sc

DETAIL = {sc.DT_PROJECTS, sc.DT_LEADS, sc.DT_INCOME, sc.DT_PART, sc.DT_TASKS, sc.DT_EXPENSES}


def test_structure_rtl(ctx=None):
    ctx = ctx or get_ctx()
    for p in (harness.Q1, harness.APR):
        wb = load_workbook(ctx.paths[p]["xlsx"])
        check(all(ws.sheet_view.rightToLeft for ws in wb.worksheets), "%s: a sheet is not RTL" % p)


def test_structure_chart_datalabels(ctx=None):
    ctx = ctx or get_ctx()
    for p in (harness.Q1, harness.APR):
        with zipfile.ZipFile(ctx.paths[p]["xlsx"]) as z:
            charts = [n for n in z.namelist() if re.match(r"xl/charts/chart\d+\.xml", n)]
            check(charts, "%s: no charts found" % p)
            for n in charts:
                xml = z.read(n).decode("utf-8", "ignore")
                check(re.search(r'showVal val="1"', xml), "%s: %s has no data labels" % (p, n))


def test_structure_detail_freeze_autofilter(ctx=None):
    ctx = ctx or get_ctx()
    wb = load_workbook(ctx.paths[harness.Q1]["xlsx"])
    nodata = _nfc(sc.S_NODATA)
    seen = set()
    for ws in wb.worksheets:
        if ws.title in DETAIL:
            seen.add(ws.title)
            check(ws.freeze_panes, "%s: no frozen header" % ws.title)
            a4 = ws.cell(4, 1).value
            has_data = a4 is not None and _nfc(a4) != nodata
            if has_data:
                check(ws.auto_filter.ref, "%s: no autofilter" % ws.title)
    check(seen == DETAIL, "missing detail sheets: %r" % (DETAIL - seen))


def test_structure_two_income_pictures(ctx=None):
    ctx = ctx or get_ctx()
    c = ctx.reports[harness.Q1]["cash"]
    check(c["net_total"] != c["gross_total"] and c["net_total"] > 0 and c["gross_total"] > 0,
          "two income pictures are not distinct")


def test_structure_no_pdf(ctx=None):
    ctx = ctx or get_ctx()
    for root, _dirs, files in os.walk(os.path.join(ROOT, "outputs")):
        if "_test" in root:
            continue
        for f in files:
            check(not f.lower().endswith(".pdf"), "pdf deliverable present: %s" % os.path.join(root, f))


def test_structure_outputs_clean(ctx=None):
    """outputs/ is organized year / type / period; every leaf period folder holds exactly
    one xlsx + one html and nothing else (no loose files at any level, valid names)."""
    ctx = ctx or get_ctx()
    base = os.path.join(ROOT, "outputs")
    types = {_nfc(t) for t in (sc.TYPE_MONTHLY, sc.TYPE_QUARTERLY, sc.TYPE_ANNUAL)}
    for year in os.listdir(base):
        if year.startswith("_"):           # scratch (e.g. _test) is not a deliverable
            continue
        ypath = os.path.join(base, year)
        check(os.path.isdir(ypath), "loose file directly under outputs/: %s" % year)
        check(re.fullmatch(r"\d{4}", year), "unexpected top-level under outputs/: %s" % year)
        for typ in os.listdir(ypath):
            tpath = os.path.join(ypath, typ)
            check(os.path.isdir(tpath), "loose file under outputs/%s/: %s" % (year, typ))
            check(_nfc(typ) in types, "unexpected type folder under outputs/%s/: %s" % (year, typ))
            for period in os.listdir(tpath):
                ppath = os.path.join(tpath, period)
                check(os.path.isdir(ppath), "loose file under outputs/%s/%s/: %s" % (year, typ, period))
                check(re.fullmatch(r"\d{4}-\d{2}|\d{4}-Q\d|\d{4}", period),
                      "unexpected period folder: %s" % period)
                files = os.listdir(ppath)
                xlsx = [f for f in files if f.endswith(".xlsx")]
                htmls = [f for f in files if f.endswith(".html")]
                check(len(xlsx) == 1 and len(htmls) == 1 and len(files) == 2,
                      "%s: expected exactly 1 xlsx + 1 html, got %r" % (period, files))


def test_structure_html_svg(ctx=None):
    ctx = ctx or get_ctx()
    for p in (harness.Q1, harness.APR):
        with open(ctx.paths[p]["html"], encoding="utf-8") as fh:
            t = fh.read()
        check("<svg" in t, "%s: html has no inline svg" % p)
        check(re.search(r"<text[^>]*>[^<]*[\d,]+", t), "%s: svg has no numeric data labels" % p)
