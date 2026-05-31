# -*- coding: utf-8 -*-
"""Same period built twice -> identical xlsx cell values (skipping the one date cell)
and identical HTML after stripping the generation-date stamp."""
import os

import harness
from harness import check, get_ctx, TMP, strip_gen, html_text, _nfc
from scripts import generator, preview
from scripts import style_config as sc
from openpyxl import load_workbook


def _build_pair(ctx, p):
    rep = ctx.reports[p]
    ax, bx = os.path.join(TMP, p + "_a.xlsx"), os.path.join(TMP, p + "_b.xlsx")
    ah, bh = os.path.join(TMP, p + "_a.html"), os.path.join(TMP, p + "_b.html")
    generator.build_workbook(rep, ax, ctx.data)
    generator.build_workbook(rep, bx, ctx.data)
    preview.build_html(rep, ctx.data, ah)
    preview.build_html(rep, ctx.data, bh)
    return ax, bx, ah, bh


def _gen_coord(ws):
    """The only date-bearing overview cell — the generation stamp (S_GENERATED prefix)."""
    pref = _nfc(sc.S_GENERATED)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and _nfc(cell.value).startswith(pref):
                return cell.coordinate
    return None


def _check_xlsx(ctx, p):
    ax, bx, _, _ = _build_pair(ctx, p)
    wa, wb = load_workbook(ax), load_workbook(bx)
    check([s.title for s in wa.worksheets] == [s.title for s in wb.worksheets],
          "%s sheet set differs" % p)
    for sa, sb in zip(wa.worksheets, wb.worksheets):
        skip = _gen_coord(sa) if sa.title == sc.T_OVERVIEW else None
        for ra, rb in zip(sa.iter_rows(), sb.iter_rows()):
            for ca, cb in zip(ra, rb):
                if skip and ca.coordinate == skip:
                    continue
                check(ca.value == cb.value,
                      "%s %s!%s: %r != %r" % (p, sa.title, ca.coordinate, ca.value, cb.value))


def _check_html(ctx, p):
    _, _, ah, bh = _build_pair(ctx, p)
    check(strip_gen(html_text(ah)) == strip_gen(html_text(bh)),
          "%s html differs after stripping the generation stamp" % p)


def test_determinism_xlsx_q1(ctx=None):
    _check_xlsx(ctx or get_ctx(), harness.Q1)


def test_determinism_xlsx_april(ctx=None):
    _check_xlsx(ctx or get_ctx(), harness.APR)


def test_determinism_html_q1(ctx=None):
    _check_html(ctx or get_ctx(), harness.Q1)


def test_determinism_html_april(ctx=None):
    _check_html(ctx or get_ctx(), harness.APR)
