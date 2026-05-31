# -*- coding: utf-8 -*-
"""Shared test infrastructure.

Builds the artifacts ONCE (loader -> metrics -> generator/preview into outputs/_test/)
and exposes value extractors for the xlsx and html so the parity / reconciliation /
determinism / structure tests can read what was actually rendered. The build functions
are pure (no PDF side effect, unlike run.main), so building into a scratch folder is
safe and does not disturb the shipped deliverables.
"""
import json
import os
import re
import sys

ROOT = os.path.dirname(os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
if ROOT not in sys.path:
    sys.path.insert(0, ROOT)

from openpyxl import load_workbook  # noqa: E402

from scripts import loader, metrics, validator, targets, generator, preview, run  # noqa: E402
from scripts import viewmodel as vm  # noqa: E402

INPUTS = os.path.join(ROOT, "inputs")
TMP = os.path.join(ROOT, "outputs", "_test")          # excluded from the clean-folder check
BASELINE = os.path.join(os.path.dirname(os.path.abspath(__file__)), "baseline.json")
Q1, APR = "2026-Q1", "2026-04"
MONTHS3 = ("2026-01", "2026-02", "2026-03")
EPS = 1e-6


def check(cond, msg):
    """Assertion that the runner reports as a FAIL with `msg` as the detail."""
    if not cond:
        raise AssertionError(msg)


def figures(rep):
    """Canonical key figures of a report (same keys as baseline.json)."""
    return {
        "leads": rep["leads"]["total"],
        "deals": rep["pipeline"]["deals_total"],
        "pipeline_value": rep["pipeline"]["value_total"],
        "net": rep["cash"]["net_total"],
        "gross": rep["cash"]["gross_total"],
        "payments": rep["cash"]["payments_total"],
        "expenses_base": rep["expenses"]["total"],
        "expenses_period": rep["expenses"]["total"] * len(rep["months"]),
        "profit": rep["profit"]["total"],
        "tasks": rep["tasks"]["total"],
        "conversion": rep["cohort"]["headline_rate"],
        "attr_matched": rep["attribution"]["matched"],
        "attr_total": rep["attribution"]["total"],
    }


def full_report(data, report_load, period):
    """Replicate run.main's report enrichment (missing/targets/prev) WITHOUT side
    effects (no blank-quarter Word emission, no PDF)."""
    rep = metrics.compute(data, period)
    _blocking, missing = validator.validate_inputs(data, report_load)
    rep["missing"] = missing + rep["missing"]
    rep["targets"] = targets.find_targets(INPUTS, period)
    rep["prev"] = run._prev_kpis(data, rep)
    return rep


class Ctx(object):
    """Built-once fixture shared across all tests."""

    def __init__(self):
        if os.path.isdir(TMP):
            for f in os.listdir(TMP):
                try:
                    os.remove(os.path.join(TMP, f))
                except OSError:
                    pass
        os.makedirs(TMP, exist_ok=True)
        self.data, self.report_load = loader.load_all()
        self.reports = {p: full_report(self.data, self.report_load, p)
                        for p in (Q1, APR) + MONTHS3}
        self.paths = {}
        for p in (Q1, APR):
            xlsx = os.path.join(TMP, p + ".xlsx")
            htmlp = os.path.join(TMP, p + ".html")
            generator.build_workbook(self.reports[p], xlsx, self.data)
            preview.build_html(self.reports[p], self.data, htmlp)
            self.paths[p] = {"xlsx": xlsx, "html": htmlp}
        with open(BASELINE, encoding="utf-8") as fh:
            self.baseline = json.load(fh)


_CTX = None


def get_ctx():
    global _CTX
    if _CTX is None:
        _CTX = Ctx()
    return _CTX


# ----------------------------------------------------------------- extractors
def _nfc(s):
    return metrics.nfc(s).strip() if s is not None else ""


def load_ws(xlsx, sheet_name):
    return load_workbook(xlsx)[sheet_name]


def find_label_cell(ws, text):
    """First cell whose NFC value equals `text`; returns (row, col) or None."""
    target = _nfc(text)
    for row in ws.iter_rows():
        for cell in row:
            if cell.value is not None and _nfc(cell.value) == target:
                return cell.row, cell.column
    return None


def find_header_col(ws, hdr_row, substr):
    sub = _nfc(substr)
    for cell in ws[hdr_row]:
        if cell.value is not None and sub in _nfc(cell.value):
            return cell.column
    return None


def matrix_summary_value(ws, label, n_vcols):
    """Read a matrix row's summary cell (the last value column). For a single-month
    period (n_vcols==1) that last column IS the month, which equals the total."""
    pos = find_label_cell(ws, label)
    if pos is None:
        return None
    r, c = pos
    return ws.cell(r, c + n_vcols).value


def kpi_value(ws, label):
    """KPI card value sits one row above its label (same column)."""
    pos = find_label_cell(ws, label)
    if pos is None:
        return None
    r, c = pos
    return ws.cell(r - 1, c).value


def html_text(path):
    with open(path, encoding="utf-8") as fh:
        return fh.read()


def strip_gen(s):
    """Blank the generation-date stamp so determinism compares everything else."""
    return re.sub(r'<div class="gen">.*?</div>', '<div class="gen"></div>', s, flags=re.S)


def html_fig(htxt, fig):
    """Exact value rendered in a KPI card carrying data-fig=`fig`."""
    m = re.search(r'data-fig="%s"[^>]*>([^<]+)<' % re.escape(fig), htxt)
    return m.group(1) if m else None
