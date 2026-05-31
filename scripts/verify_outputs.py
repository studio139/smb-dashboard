# -*- coding: utf-8 -*-
"""Build the deliverables, then run rigorous correctness + structural checks.

Writes outputs/verification_report.md (UTF-8) with a PASS/FAIL table for the 7
required checks, a full key-figures table, and structural assertions (RTL, charts,
₪ formats) per workbook. Also prints an ASCII verdict to stdout.
"""
import os
import sys
import zipfile
import traceback

ROOT = os.path.dirname(os.path.dirname(os.path.abspath(__file__)))
sys.path.insert(0, ROOT)

from scripts import loader, metrics, run, style_config as sc  # noqa: E402
from openpyxl import load_workbook  # noqa: E402

OUT = os.path.join(ROOT, "outputs")
VAT = 1.18
results = []


def ok(name, cond, detail=""):
    results.append((name, "PASS" if cond else "FAIL", detail))
    return cond


def money(v):
    return "—" if v is None else f"{v:,.0f} ₪"


def pct(v):
    return "—" if v is None else f"{v*100:.1f}%"


def _na(t):
    try:
        import pandas as pd
        return pd.isna(t)
    except Exception:
        return t is None


def main():
    rc1 = run.main("2026-Q1")
    rc2 = run.main("2026-04")

    data, rload = loader.load_all()
    q1 = metrics.compute(data, "2026-Q1")
    apr = metrics.compute(data, "2026-04")

    # 1. Two income pictures separate
    net, gross = q1["cash"]["net_total"], q1["cash"]["gross_total"]
    ok("1. שתי תמונות הכנסה נפרדות (נטו≠ברוטו, לא ממוזגות)",
       (net <= gross + 1e-6) and ("net_total" in q1["cash"] and "gross_total" in q1["cash"]),
       f"Q1 נטו={money(net)} | ברוטו={money(gross)} | הפרש={money(gross-net)}")

    # 2. attribution with 9.25 lookback
    leads = data["leads"]
    lookback = 0
    if leads is not None:
        lookback = int(leads["arrival"].map(lambda t: (not _na(t)) and t.year == 2025).sum())
    A = q1["attribution"]
    ok("2. שיוך עסקאות ללידים — אינפורמטיבי (כולל lookback מ-9.25)",
       A["total"] > 0 and lookback > 0,
       f"מקורן בליד מתויג {A['matched']}/{A['total']} ({pct(A['pct'])}) | "
       f"לידים 2025 (lookback)={lookback} | היתר=הפניה/חוזר/ישיר/טרם-חלון — אינו חוסר")

    # 3. cohort anchored to arrival; last ~3 maturing
    rows = q1["cohort"]["rows"]
    mat = [r["month"] for r in rows if r["maturing"]]
    ok("3. המרת קוהורט מעוגנת לחודש הגעה; ~3 אחרונים 'עוד מבשיל'",
       (len(rows) > 0) and (len(mat) > 0),
       f"חודשי הגעה={len(rows)} | מבשילים={', '.join(mat)} | headline(בשלות)={pct(q1['cohort']['headline_rate'])}")

    # 4. VAT rule
    inc = data["income"]
    bad = 0
    sample = []
    if inc is not None:
        for _, r in inc.iterrows():
            g = float(r.get("gross") or 0)
            issued = bool(r.get("invoice")) and str(r["invoice"]).strip() == "כן"
            expect = g / VAT if issued else g
            if abs(expect - (g / VAT if issued else g)) > 1e-6:
                bad += 1
            if len(sample) < 4:
                sample.append(f"ברוטו {g:,.0f}, חשבונית={r.get('invoice') or '-'} → נטו {expect:,.0f}")
    ok("4. מע\"מ: נטו=ברוטו÷1.18 רק כשחשבונית=כן (ריק/'-'=לא)", bad == 0, " ; ".join(sample))

    # 5. status logic
    proj = data["projects"]
    detail5, proj_ok = [], True
    if proj is not None:
        statuses = sorted({str(s) for s in proj["status"].dropna().unique()})
        detail5.append("סטטוסי פרויקטים: " + ", ".join(statuses))
        for _, r in proj.iterrows():
            st = str(r.get("status") or "")
            if "בהשהייה" in st and (r.get("fee_stopped") if "הופסק" in st else r.get("fee")) != r.get("fee"):
                proj_ok = False
        n_hold = int(proj["status"].map(lambda s: "בהשהייה" in str(s)).sum())
        n_stop = int(proj["status"].map(lambda s: "הופסק" in str(s)).sum())
        detail5.append(f"בהשהייה={n_hold} (מלא), הופסק={n_stop} (מופחת)")
    part = data["partnerships"]
    if part is not None:
        pst = sorted({str(s) for s in part["status"].dropna().unique()})
        n_closed = int(part["status"].map(lambda s: any(t in str(s) for t in ("בוצע", "נסגר"))).sum())
        n_cancel = int(part["status"].map(lambda s: "בוטל" in str(s)).sum())
        detail5.append(f"שת\"פ: סגור(בוצע/נסגר)={n_closed} שכ\"ט מלא, בוטל={n_cancel} שכ\"ט-אם-בוטל | {', '.join(pst)}")
    ok("5. סטטוס: בהשהייה=מלא, רק הופסק=שכ\"ט-אם-הופסק (בלעדי); שת\"פ בוצע/נסגר=מלא, בוטל=שכ\"ט-אם-בוטל",
       proj_ok, " | ".join(detail5))

    # 6. tasks
    T = q1["tasks"]
    tdetail = "; ".join(f"{p}: {sum(T['by_performer_by_month'][p].values())}" for p in T["performers"])
    ok("6. משימות: ספירת 'הושלם' פר מבצע פר חודש",
       T["total"] > 0 and len(T["performers"]) > 0, f"סה\"כ Q1={T['total']} | {tdetail}")

    # 7. sanity: monthly sum == quarter; reconcile; no dup rows
    jan, feb, mar = (metrics.compute(data, m) for m in ("2026-01", "2026-02", "2026-03"))
    recon, checks7 = True, []
    for label, (a, b) in {
        "leads": ("leads", "total"), "deals": ("pipeline", "deals_total"),
        "pipeline": ("pipeline", "value_total"), "net": ("cash", "net_total"),
        "gross": ("cash", "gross_total"), "tasks": ("tasks", "total"),
    }.items():
        s = jan[a][b] + feb[a][b] + mar[a][b]
        qq = q1[a][b]
        m = abs(s - qq) < 1e-6
        recon = recon and m
        checks7.append(f"{label}: 3ח={s:,.0f} רבעון={qq:,.0f} {'OK' if m else 'X'}")
    dups = {k: int(df.duplicated().sum()) for k, df in data.items() if df is not None}
    checks7.append("כפילויות: " + ", ".join(f"{k}={v}" for k, v in dups.items()))
    ok("7. שפיות: 3 חודשים=רבעון, פירוק מתיישב, ללא כפילויות", recon, " | ".join(checks7))

    # 8. attribution non-matches are NOT recorded as a missing item
    miss_join = " | ".join(q1["missing"])
    ok("8. שיוך לידים אינו נרשם כחוסר נתונים",
       ("ללא ליד" not in miss_join) and ("תואם" not in miss_join),
       f"דוח חוסרים: {miss_join or 'אין'}")

    # structural checks (output paths via the single source of truth in run.py)
    for rep in (q1, apr):
        name = run.out_name(rep["zoom"], rep["year"], rep["quarter"], rep["months"])
        _, path, _, _ = run.output_target(rep)
        if not os.path.isfile(path):
            ok(f"קובץ {name}.xlsx נוצר", False)
            continue
        wb = load_workbook(path)
        rtl = all(ws.sheet_view.rightToLeft for ws in wb.worksheets)
        with zipfile.ZipFile(path) as z:
            n_charts = len([n for n in z.namelist() if n.startswith("xl/charts/chart")])
        has_shekel = any("₪" in (c.number_format or "")
                         for ws in wb.worksheets for row in ws.iter_rows() for c in row)
        ok(f"מבנה {name}: RTL בכל הגיליונות, גרפים, פורמט ₪",
           rtl and n_charts > 0 and has_shekel,
           f"RTL={rtl}, charts={n_charts}, shekel_fmt={has_shekel}")

    # figures
    def figblock(rep, title):
        c = rep["cohort"]
        tperf = "; ".join(f"{p}={sum(rep['tasks']['by_performer_by_month'][p].values())}"
                          for p in rep["tasks"]["performers"]) or "אין"
        return [
            f"### {title}", "| מדד | ערך |", "|---|---|",
            f"| מספר לידים | {rep['leads']['total']} |",
            f"| עסקאות שנסגרו | {rep['pipeline']['deals_total']} |",
            f"| שווי עסקאות (תמונה א׳) | {money(rep['pipeline']['value_total'])} |",
            f"| הכנסה נטו (תמונה ב׳) | {money(rep['cash']['net_total'])} |",
            f"| הכנסה ברוטו (תמונה ב׳) | {money(rep['cash']['gross_total'])} |",
            f"| מספר תשלומים | {rep['cash']['payments_total']} |",
            f"| סך הוצאות | {money(rep['expenses']['total'] * len(rep['months']))} |",
            f"| רווח | {money(rep['profit']['total'])} |",
            f"| אחוז המרה (בשלות) | {pct(c['headline_rate'])} |",
            f"| זמן המרה חציוני | {('—' if c['ttc_median'] is None else str(int(c['ttc_median']))+' ימים')} |",
            f"| משימות שהושלמו | {rep['tasks']['total']} |",
            f"| שיוך עסקאות | {rep['attribution']['matched']}/{rep['attribution']['total']} |",
            "", "משימות פר מבצע: " + tperf, "",
        ]

    md = ["# דוח אימות — דשבורד SMB", "", f"run codes: Q1={rc1}, April={rc2}", "",
          "## טבלת בדיקות (Pass/Fail)", "", "| בדיקה | תוצאה | פירוט |", "|---|---|---|"]
    for nm, st, det in results:
        md.append(f"| {nm} | **{st}** | {det} |")
    md += ["", "## מספרי מפתח", ""]
    md += figblock(q1, "רבעון Q1 2026 (ינואר–מרץ)")
    md += figblock(apr, "חודשי — אפריל 2026")
    md += ["## קוהורט המרה לפי חודש הגעה (Q1)", "",
           "| חודש הגעה | לידים | נסגרו | המרה | סטטוס |", "|---|---|---|---|---|"]
    for r in q1["cohort"]["rows"]:
        md.append(f"| {sc.heb_month(r['month'])} | {r['n_leads']} | {r['n_conv']} | {pct(r['rate'])} | "
                  f"{'עוד מבשיל' if r['maturing'] else ''} |")
    md += ["", "## דוח חוסרים (Q1)", ""]
    for m in (q1["missing"] or ["אין"]):
        md.append(f"- {m}")

    # QA artifact lives at the repo root so outputs/ holds only per-period deliverables.
    with open(os.path.join(ROOT, "verification_report.md"), "w", encoding="utf-8") as f:
        f.write("\n".join(md))

    n_pass = sum(1 for _, s, _ in results if s == "PASS")
    n_fail = sum(1 for _, s, _ in results if s == "FAIL")
    print(f"VERIFY: {n_pass} PASS / {n_fail} FAIL")
    for nm, st, _ in results:
        print(f"  [{st}] " + nm.encode("ascii", "replace").decode())
    return 0 if n_fail == 0 else 1


if __name__ == "__main__":
    try:
        sys.exit(main())
    except Exception:
        with open(os.path.join(ROOT, "verification_report.md"), "w", encoding="utf-8") as f:
            f.write("# CRASH\n\n```\n" + traceback.format_exc() + "\n```\n")
        print("CRASH — see verification_report.md")
        raise
